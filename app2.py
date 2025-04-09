# Importações 
import streamlit as st
import requests
from PIL import Image
from io import BytesIO
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from PyPDF2 import PdfReader, PdfWriter
import fitz  # PyMuPDF

# Configuração do Streamlit
st.set_page_config(page_title="Imagens para PDF", layout="wide")
st.title("📎 Ferramenta Spot - VExpenses")

# Função para extrair os dados da planilha
def extrair_links_e_ids(file):
    wb = load_workbook(file, data_only=True)
    ws = wb.active
    headers = {cell.value: idx for idx, cell in enumerate(ws[1])}

    if "Link do Anexo" not in headers or "ID da Despesa" not in headers or "ID do Relatório" not in headers:
        raise ValueError("A planilha deve conter as colunas 'Link do Anexo', 'ID da Despesa' e 'ID do Relatório'.")

    col_link = headers["Link do Anexo"]
    col_id_despesa = headers["ID da Despesa"]
    col_id_relatorio = headers["ID do Relatório"]

    dados = []
    for row in ws.iter_rows(min_row=2):
        linha_excel = row[0].row
        id_despesa = row[col_id_despesa].value
        id_relatorio = row[col_id_relatorio].value
        cell_link = row[col_link]
        url = cell_link.hyperlink.target if cell_link.hyperlink else None
        dados.append((linha_excel, id_despesa, id_relatorio, url))

    return dados

# Converte PDF para imagens usando PyMuPDF
def pdf_para_imagens(pdf_bytes):
    imagens = []
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        for page in doc:
            pix = page.get_pixmap(dpi=200)
            img = Image.open(BytesIO(pix.tobytes("png")))
            imagens.append(img)
    return imagens

# Adiciona número de página
def adicionar_numero_pagina(c, numero):
    c.setFont("Helvetica", 9)
    c.drawRightString(550, 20, f"Página {numero}")

# Upload do arquivo Excel
uploaded_file = st.file_uploader("📂 Envie a planilha (.xlsx) com os links de imagem", type=["xlsx"])
manual_uploads = {}

if uploaded_file:
    try:
        info_links = extrair_links_e_ids(uploaded_file)

        if not info_links:
            st.error("❌ Nenhum link encontrado na planilha.")
        else:
            st.success(f"✅ {len(info_links)} registros encontrados.")
            st.header("📸 Upload manual para anexos ausentes")

            for linha, id_despesa, id_relatorio, url in info_links:
                if not url:
                    st.warning(f"🔍 Imagem ausente: ID da Despesa {id_despesa} | ID do Relatório {id_relatorio}")
                    img = st.file_uploader(
                        f"Envie a imagem para linha {linha}",
                        type=["jpg", "png"],
                        key=f"upload_linha_{linha}_despesa_{id_despesa}"
                    )
                    manual_uploads[linha] = img

            imagens_pendentes = [linha for linha, id_d, id_r, url in info_links if not url and not manual_uploads.get(linha)]
            if imagens_pendentes:
                st.info("⏳ Aguardando envio de todas as imagens manuais antes de gerar o PDF.")
                st.stop()

            if st.button("🔄 Gerar PDF"):
                erros = []
                pdf_writer = PdfWriter()
                pagina_atual = 1

                with st.spinner("⏳ Gerando PDF em ordem..."):
                    for linha, id_despesa, id_relatorio, url in info_links:
                        try:
                            st.write(f"🔄 Linha {linha} | ID Despesa: {id_despesa} | ID Relatório: {id_relatorio}")

                            # Baixa ou pega imagem
                            if not url:
                                uploaded_image = manual_uploads.get(linha)
                                if not uploaded_image:
                                    raise ValueError("Imagem manual não enviada.")
                                img = Image.open(uploaded_image).convert("RGB")
                                imagens = [img]
                                st.write("✅ Imagem manual carregada.")
                            else:
                                if not url.startswith("http"):
                                    url = "https://" + url

                                st.write(f"🌐 Buscando URL: {url}")
                                response = requests.get(url, timeout=20)
                                st.write(f"📡 Status: {response.status_code}")
                                content_type = response.headers.get('Content-Type', '')
                                st.write(f"📄 Content-Type: {content_type}")
                                response.raise_for_status()

                                if 'pdf' in content_type:
                                    imagens = pdf_para_imagens(response.content)
                                    st.write(f"🖼️ PDF com {len(imagens)} página(s) convertidas em imagem.")
                                else:
                                    img = Image.open(BytesIO(response.content)).convert("RGB")
                                    extrema = img.getextrema()

                                    if all(e[0] == e[1] for e in extrema):
                                        st.write("⚠️ Imagem possivelmente em branco.")
                                        raise ValueError("Imagem aparentemente em branco (extremos iguais).")

                                    imagens = [img]
                                    st.write("✅ Imagem carregada com sucesso.")

                            # Página combinando texto e imagem(s)
                            img_pdf_buffer = BytesIO()
                            c = canvas.Canvas(img_pdf_buffer, pagesize=letter)

                            texto_ids = f"ID da Despesa: {id_despesa} / ID do Relatório: {id_relatorio}"
                            c.setFont("Helvetica-Bold", 12)
                            c.drawString(50, 750, texto_ids)
                            adicionar_numero_pagina(c, pagina_atual)
                            pagina_atual += 1

                            y_pos = 630

                            for img in imagens:
                                page_width, page_height = letter
                                max_width, max_height = 500, 600
                                img_width, img_height = img.size
                                scale = min(max_width / img_width, max_height / img_height)
                                new_width = img_width * scale
                                new_height = img_height * scale
                                x_pos = (page_width - new_width) / 2

                                img_io = BytesIO()
                                img.save(img_io, format='PNG')
                                img_io.seek(0)
                                c.drawImage(ImageReader(img_io), x_pos, y_pos - new_height, width=new_width, height=new_height)
                                y_pos -= new_height + 20

                            adicionar_numero_pagina(c, pagina_atual - 1)
                            c.showPage()
                            c.save()

                            img_pdf_buffer.seek(0)
                            img_reader = PdfReader(img_pdf_buffer)
                            for page in img_reader.pages:
                                pdf_writer.add_page(page)

                        except Exception as e:
                            erros.append((linha, id_despesa, id_relatorio, e))
                            st.warning(f"Erro na linha {linha} | ID Despesa: {id_despesa}, Relatório: {id_relatorio} → {e}")

                            error_pdf_buffer = BytesIO()
                            c = canvas.Canvas(error_pdf_buffer, pagesize=letter)
                            texto = f"ID da Despesa: {id_despesa} / ID do Relatório: {id_relatorio}"
                            c.setFont("Helvetica", 10)
                            c.drawString(50, 750, texto)
                            c.setFont("Helvetica-Bold", 12)
                            c.setFillColorRGB(1, 0, 0)
                            c.drawString(50, 700, "⚠️ Imagem não pôde ser carregada.")
                            adicionar_numero_pagina(c, pagina_atual)
                            pagina_atual += 1
                            c.showPage()
                            c.save()
                            error_pdf_buffer.seek(0)
                            error_page = PdfReader(error_pdf_buffer)
                            pdf_writer.add_page(error_page.pages[0])

                if pdf_writer.pages:
                    final_output = BytesIO()
                    pdf_writer.write(final_output)
                    final_output.seek(0)

                    st.success("✅ PDF gerado com sucesso!")
                    st.download_button(
                        label="📥 Baixar PDF",
                        data=final_output,
                        file_name="anexos_ordenados.pdf",
                        mime="application/pdf"
                    )

                    if erros:
                        st.markdown("### ❌ Falhas detectadas")
                        falhas_links = []
                        for linha, id_despesa, id_relatorio, erro in erros:
                            texto_erro = f"Linha {linha} | Despesa: {id_despesa} | Relatório: {id_relatorio}"
                            st.write(texto_erro)
                            falhas_links.append(texto_erro)

                        texto_para_copiar = "\n".join(falhas_links)
                        st.download_button(
                            label="📄 Baixar lista de falhas (.txt)",
                            data=texto_para_copiar,
                            file_name="falhas_links.txt",
                            mime="text/plain"
                        )
                else:
                    st.error("❌ Nenhuma página foi gerada.")
    except Exception as e:
        st.error(f"Erro ao processar o arquivo: {e}")
